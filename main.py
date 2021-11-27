#garpozir@gmail.com

import time as t
import time,shutil
from pandas.plotting import table
import matplotlib.pyplot as plt,mpld3
from openpyxl.chart import DoughnutChart, Reference
from openpyxl.chart.series import DataPoint
import pandas as pd,os
from os import system,name
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Color,Alignment,colors,PatternFill
import numpy as np,sys,re
from mpl_toolkits.mplot3d import Axes3D

if name == 'nt':
     _ = system('cls')
else:
    _ = system('clear')

if os.path.isfile('in.xlsx'):shutil.copyfile('in.xlsx','do.xlsx')
else:exit()

def Cube(input_array, oupt_array, color_input, color_output, axis_size):

    if len(oupt_array) > 0:
        fig = plt.figure()

        r = [0, 1]
        X, Y = np.meshgrid(r, r)
        one = np.ones(4).reshape(2, 2)

        for point in input_array:
            ax = fig.gca(projection='3d')
            ax.set_xlim(-axis_size, axis_size)
            ax.set_ylim(-axis_size, axis_size)
            ax.set_zlim(-axis_size, axis_size)
            x = point[0]
            y = point[1]
            z = point[2]

            ax.plot_surface(X + one * (x - 0.5), Y + one * (y - 0.5), one * (z + 0.5), color=color_input)
            ax.plot_surface(X + one * (x - 0.5), Y + one * (y - 0.5), one * (z - 0.5), color=color_input)
            ax.plot_surface(X + one * (x - 0.5), one * (y + 0.5), Y + one * (z - 0.5), color=color_input)
            ax.plot_surface(X + one * (x - 0.5), one * (y - 0.5), Y + one * (z - 0.5), color=color_input)
            ax.plot_surface(one * (x + 0.5), X + one * (y - 0.5), Y + one * (z - 0.5), color=color_input)
            ax.plot_surface(one * (x - 0.5), X + one * (y - 0.5), Y + one * (z - 0.5), color=color_input)

        for point in oupt_array:
            ax = fig.gca(projection='3d')
            ax.set_xlim(-axis_size, axis_size)
            ax.set_ylim(-axis_size, axis_size)
            ax.set_zlim(-axis_size, axis_size)
            x = point[0]
            y = point[1]
            z = point[2]

            ax.plot_surface(X + one * (x - 0.5), Y + one * (y - 0.5), one * (z + 0.5), color=color_output)
            ax.plot_surface(X + one * (x - 0.5), Y + one * (y - 0.5), one * (z - 0.5), color=color_output)
            ax.plot_surface(X + one * (x - 0.5), one * (y + 0.5), Y + one * (z - 0.5), color=color_output)
            ax.plot_surface(X + one * (x - 0.5), one * (y - 0.5), Y + one * (z - 0.5), color=color_output)
            ax.plot_surface(one * (x + 0.5), X + one * (y - 0.5), Y + one * (z - 0.5), color=color_output)
            ax.plot_surface(one * (x - 0.5), X + one * (y - 0.5), Y + one * (z - 0.5), color=color_output)

        ax.set_xlabel('X')
        ax.set_ylabel('Y')
        ax.set_zlabel('Z')

    else:
        print('final array is empty!')

def GetPlotAxisSize(input_array, output_array):
    if len(output_array) > 0:
        m1 = np.max(input_array[:, :-1])
        m2 = np.max(output_array[:, :-1])
        return int(max(m1, m2)) * 2
    else:
        print('final array is empty!')
        return False

start = t.time()

oo='in'
wbi = openpyxl.load_workbook(f'{oo}.xlsx')
sheeti = wbi['Sheet1']
coul=[]
for row in ['A','B','C','D','E']:
    coul.append(sheeti[f"{row}1"].value)
    if row=='B' or row=='D' or row=='E':sheeti.column_dimensions[f'{row}'].width=30
    for col in range(1,5):
        if row=='A':sheeti.row_dimensions[int(col)].height=20
        sheeti[f'{row}{col}'].alignment = Alignment(horizontal='center',vertical='center')
        if col%2==0:
            sheeti[f'{row}{col}'].fill = PatternFill(fgColor="E6E6E6", fill_type = "solid")
sheeti.cell(row = col+2, column = 2).alignment = Alignment(horizontal='left',vertical='center')
wbi.close()
df=pd.read_excel(f'{oo}.xlsx')
df=pd.DataFrame(df,)
a=df.values.tolist()
a=np.array(a)

p = int(input('p '))
cr =int(input('cr '))
er =int(input('er '))
rr =int(input('rr '))
y = float(input('y '))
tr =int(input('tr '))

a[:, 3] = ((p - rr) * a[:, 3] * y - (er + cr)) * tr

xmax = int(max(a[:, 0]))
ymax = int(max(a[:, 1]))
zmax = int(max(a[:, 2]))
xmin = int(min(a[:, 0]))
ymin = int(min(a[:, 1]))
zmin = int(min(a[:, 2]))
xs = 2
ys = 2
zs = 2
nx = round((xmax - xmin) / xs)
ny = round((ymax - ymin) / ys)
nz = round((zmax - zmin) / zs)
sx = 3
sy = 2
sz = 2

b = []
counter1 = 0
counter2 = 0
for i in range(xmin, xmax + xs, xs):
    for j in range(ymin, ymax + ys, ys):
        for k in range(zmin, zmax + zs, zs):
            b.append([i, j, k])
            counter1 += 1

for i in a:
    for t, j in enumerate(b):
        if list(i[:3]) == j:
            b[t].append(list(i)[-1])

for i in b:
    if len(i) == 3:
        i.append(0)

b = np.array(b)

reindexed_b = []
for j in range(ymin, ymax + ys, ys):
    for k in range(zmin, zmax + zs, zs):
        for i in range(xmin, xmax + xs, xs):
            reindexed_b.append([i, j, k])

for i, item in enumerate(reindexed_b):
    for it in b:
        if item == list(it)[:-1]:
            reindexed_b[i].append(it[-1])
            break

final = []
i = 0

while True:

    Sum = 0
    for l1 in range(sy):
        for l2 in range(sz):
            for l3 in range(sx):
                Sum += reindexed_b[(int(nx) + 1) * l2 + (int(nx) + 1) * (int(nz) + 1) * l1 + l3 + i][3]

    if Sum > 0:
        for l1 in range(sy):
            for l2 in range(sz):
                for l3 in range(sx):
                    if reindexed_b[(int(nx) + 1) * l2 + (int(nx) + 1) * (int(nz) + 1) * l1 + l3 + i] not in final:
                        final.append(reindexed_b[(int(nx) + 1) * l2 + (int(nx) + 1) * (int(nz) + 1) * l1 + l3 + i])

    if (int(nx) + 1) * (sx - 1) + (int(nx) + 1) * (int(nz) + 1) * (sz - 1) + (sx - 1) + i + 1 == len(reindexed_b):
        break
    i += 1

final = np.array(final)

nrows, ncols = a.shape
dtype = {'names': ['f{}'.format(i) for i in range(ncols)],
         'formats': ncols * [a.dtype]}
C = np.intersect1d(a.view(dtype), final.view(dtype))
C = C.view(a.dtype).reshape(-1, ncols)

final = C

teedad_block_estekhraji = len(final)
majmoe_cost = sum(final[:, 3])

def se():
    majmoe=str(majmoe_cost).split('.')[1]
    cost=str(majmoe_cost).split('.')[0]
    txt,cost=cost,'{:,}'
    majmoe_cost=(cost.format(int(txt)))
    majmoe_cost=majmoe_cost+'.'+majmoe

majmoe_cost=str(majmoe_cost)
teedad_block_estekhraji=str(teedad_block_estekhraji)

df=pd.DataFrame(final)
df.index+=1
df.to_excel('out.xlsx',index_label='id')
time.sleep(2)
o='out'
wb = openpyxl.load_workbook(f'{o}.xlsx')
sheet = wb['Sheet1']
sheet.cell(row = 2, column = 7).value = f'majmoe_cost= {majmoe_cost}'
sheet.cell(row = 3, column = 7).value = f'teedad_block_estekhraji= {teedad_block_estekhraji}'
wb.save(f'{o}.xlsx')
wb.close()

axis_size = GetPlotAxisSize(a, final)
Cube(a, final, 'red', 'blue', axis_size)
plt.savefig('3D.png')

time.sleep(1)
wb = openpyxl.load_workbook(f'{o}.xlsx')
sheet = wb['Sheet1']
img = openpyxl.drawing.image.Image('3D.png')
img.anchor = 'G5'
sheet.add_image(img)
wb.save(f'{o}.xlsx')
wb.close()

wb = openpyxl.load_workbook('do.xlsx')
sheet = wb['Sheet1']
rowi='D'
for col in range(2,sheet.max_row+1):
    exl=float(sheet[f"{rowi}{col}"].value)
    bev=(((p-rr)*exl*y-(cr+er))*tr)
    sheet.cell(row = col, column = 4).value = str(bev)
wb.save('do.xlsx')
wb.close()

ca=input('please enter to exit...')
