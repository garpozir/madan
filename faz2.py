#garpozir@gmail.com

import time as t
import time
from pandas.plotting import table
import matplotlib.pyplot as plt,mpld3
from openpyxl.chart import DoughnutChart, Reference
from openpyxl.chart.series import DataPoint
import pandas as pd,os
from os import system,name
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Color,Alignment,colors,PatternFill
import numpy as np,sys,re
from mpl_toolkits.mplot3d import Axes3D

if name == 'nt':
     _ = system('cls')
else:
    _ = system('clear')

def Cube(input_array, oupt_array, color_input, color_output, axis_size):

    if len(oupt_array) > 0:
        fig = plt.figure()

        r = [0, 1]
        X, Y = np.meshgrid(r, r)
        one = np.ones(4).reshape(2, 2)

        for point in input_array:
            ax = fig.gca(projection='3d')
            ax.set_xlim(-axis_size, axis_size)
            ax.set_ylim(-axis_size, axis_size)
            ax.set_zlim(-axis_size, axis_size)
            x = point[0]
            y = point[1]
            z = point[2]

            ax.plot_surface(X + one * (x - 0.5), Y + one * (y - 0.5), one * (z + 0.5), color=color_input)
            ax.plot_surface(X + one * (x - 0.5), Y + one * (y - 0.5), one * (z - 0.5), color=color_input)
            ax.plot_surface(X + one * (x - 0.5), one * (y + 0.5), Y + one * (z - 0.5), color=color_input)
            ax.plot_surface(X + one * (x - 0.5), one * (y - 0.5), Y + one * (z - 0.5), color=color_input)
            ax.plot_surface(one * (x + 0.5), X + one * (y - 0.5), Y + one * (z - 0.5), color=color_input)
            ax.plot_surface(one * (x - 0.5), X + one * (y - 0.5), Y + one * (z - 0.5), color=color_input)

        for point in oupt_array:
            ax = fig.gca(projection='3d')
            ax.set_xlim(-axis_size, axis_size)
            ax.set_ylim(-axis_size, axis_size)
            ax.set_zlim(-axis_size, axis_size)
            x = point[0]
            y = point[1]
            z = point[2]

            ax.plot_surface(X + one * (x - 0.5), Y + one * (y - 0.5), one * (z + 0.5), color=color_output)
            ax.plot_surface(X + one * (x - 0.5), Y + one * (y - 0.5), one * (z - 0.5), color=color_output)
            ax.plot_surface(X + one * (x - 0.5), one * (y + 0.5), Y + one * (z - 0.5), color=color_output)
            ax.plot_surface(X + one * (x - 0.5), one * (y - 0.5), Y + one * (z - 0.5), color=color_output)
            ax.plot_surface(one * (x + 0.5), X + one * (y - 0.5), Y + one * (z - 0.5), color=color_output)
            ax.plot_surface(one * (x - 0.5), X + one * (y - 0.5), Y + one * (z - 0.5), color=color_output)

        ax.set_xlabel('X')
        ax.set_ylabel('Y')
        ax.set_zlabel('Z')

    else:
        print('final array is empty!')


def GetPlotAxisSize(input_array, output_array):
    if len(output_array) > 0:
        m1 = np.max(input_array[:, :-1])
        m2 = np.max(output_array[:, :-1])
        return int(max(m1, m2)) * 2
    else:
        print('final array is empty!')
        return False

start = t.time()

if os.path.isfile('in.xlsx'):pass
else:exit()
oo='in'
wbi = openpyxl.load_workbook(f'{oo}.xlsx')
sheeti = wbi['Sheet1']
coul=[]
for row in ['A','B','C','D','E']:
    coul.append(sheeti[f"{row}1"].value)
    if row=='B' or row=='D' or row=='E':sheeti.column_dimensions[f'{row}'].width=30
    for col in range(1,5):
        if row=='A':sheeti.row_dimensions[int(col)].height=20
        sheeti[f'{row}{col}'].alignment = Alignment(horizontal='center',vertical='center')
        if col%2==0:
            sheeti[f'{row}{col}'].fill = PatternFill(fgColor="E6E6E6", fill_type = "solid")
sheeti.cell(row = col+2, column = 2).alignment = Alignment(horizontal='left',vertical='center')
wbi.close()
df=pd.read_excel(f'{oo}.xlsx')
df=pd.DataFrame(df,)
a=df.values.tolist()
a=np.array(a)

p = int(input('p '))
cr =int(input('cr '))
er =int(input('er '))
rr =int(input('rr '))
y = float(input('y '))
tr =int(input('tr '))

a[:, 3] = ((p - rr) * a[:, 3] * y - (er + cr)) * tr

xmax = int(max(a[:, 0]))
ymax = int(max(a[:, 1]))
zmax = int(max(a[:, 2]))
xmin = int(min(a[:, 0]))
ymin = int(min(a[:, 1]))
zmin = int(min(a[:, 2]))
xs = 2
ys = 2
zs = 2
nx = round((xmax - xmin) / xs)

ny = round((ymax - ymin) / ys)
nz = round((zmax - zmin) / zs)
sx = 2
sy = 2
sz = 2

b = []
counter1 = 0
counter2 = 0
for i in range(xmin, xmax + xs, xs):
    for j in range(ymin, ymax + ys, ys):
        for k in range(zmin, zmax + zs, zs):
            b.append([i, j, k])
            counter1 += 1

for i in a:
    for t, j in enumerate(b):
        if list(i[:3]) == j:
            b[t].append(list(i)[-1])

for i in b:
    if len(i) == 3:
        i.append(0)

b = np.array(b)

reindexed_b = []
for j in range(ymin, ymax + ys, ys):
    for k in range(zmin, zmax + zs, zs):
        for i in range(xmin, xmax + xs, xs):
            reindexed_b.append([i, j, k])

for i, item in enumerate(reindexed_b):
    for it in b:
        if item == list(it)[:-1]:
            reindexed_b[i].append(it[-1])
            break

final = []
i = 0
extractor_worth = []
while True:

    Sum = 0
    FLAG = True
    helper = []
    max_worth = 0
    most_worth_extractor = []
    for l1 in range(sy):
        for l2 in range(sz):
            for l3 in range(sx):
                Sum += reindexed_b[(int(nx) + 1) * l2 + (int(nx) + 1) * (int(nz) + 1) * l1 + l3 + i][3]

    if Sum > 0:
        for l1 in range(sy):
            for l2 in range(sz):
                for l3 in range(sx):
                    if reindexed_b[(int(nx) + 1) * l2 + (int(nx) + 1) * (int(nz) + 1) * l1 + l3 + i] not in final:
                        helper.append(reindexed_b[(int(nx) + 1) * l2 + (int(nx) + 1) * (int(nz) + 1) * l1 + l3 + i])
                    else:
                        FLAG = False
                        helper = []
                        break
                if not FLAG:
                    break
            if not FLAG:
                break

    if FLAG:
        if Sum > max_worth:
            max_worth = Sum
            most_worth_extractor = helper


    for hlpr in helper:
        final.append(hlpr)
    if (int(nx) + 1) * (sx - 1) + (int(nx) + 1) * (int(nz) + 1) * (sz - 1) + (sx - 1) + i + 1 == len(reindexed_b):
        break
    i += 1

final = np.array(final)
most_worth_extractor = np.array(most_worth_extractor)
nrows, ncols = a.shape
dtype = {'names': ['f{}'.format(i) for i in range(ncols)],
         'formats': ncols * [a.dtype]}
C = np.intersect1d(a.view(dtype), final.view(dtype))
C = C.view(a.dtype).reshape(-1, ncols)

final = C

xmax = int(max(final[:, 0]))
ymax = int(max(final[:, 1]))
zmax = int(max(final[:, 2]))
xmin = int(min(final[:, 0]))
ymin = int(min(final[:, 1]))
zmin = int(min(final[:, 2]))

while True:
    ind = []
    for i,j in enumerate(final):
        if j[0] in [xmin,xmax] and j[3] <= 0:
            ind.append(i)

        elif j[1] in [ymin,ymax] and j[3] <= 0:
            ind.append(i)

        elif j[2] in [zmin,zmax] and j[3] <= 0:
            ind.append(i)

    if len(ind) > 0:
        final = np.delete(final,ind,0)

    else:
        break

majmoe_cost = sum(final[:, 3])
teedad_block_estekhraji = len(final)
teedad_block=len(most_worth_extractor)
oo='out_final.xlsx'
if teedad_block>0:
    df=pd.DataFrame(most_worth_extractor)
    df.index+=1
    df.to_excel(oo,index_label='id')
    time.sleep(1)
    wb = openpyxl.load_workbook(oo)
    sheet = wb['Sheet1']
    sheet.cell(row = 3, column = 7).value = f'tedad_block_estekhraji= {teedad_block_estekhraji}'
    sheet.cell(row = 4, column = 7).value = f'tedad_block= {teedad_block}'
    sheet.cell(row = 5, column = 7).value = f'majmoe_cost= {majmoe_cost}'
    sheet.cell(row = 6, column = 7).value = f'Arzesh_nahayi= {sum(most_worth_extractor[:, 3])}'
    axis_size = GetPlotAxisSize(a, final)
    Cube(a, final, 'red', 'blue', axis_size)
    plt.savefig('faz2.png')
    time.sleep(1)
    img = openpyxl.drawing.image.Image('faz2.png')
    img.anchor = 'G8'
    sheet.add_image(img)
    wb.save(oo)
    wb.close()
    if os.path.isfile('faz2.png'):
        os.remove('faz2.png')
else:
    print('teedad_block is empty!')

ca=input('please enter to exit...')
